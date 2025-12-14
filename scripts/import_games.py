#!/usr/bin/env python3
"""
Excel to SQL Import Script for 10K Scorekeeper

Usage:
    python import_games.py <excel_file> <your_user_id> [output_file]

Excel format (per sheet = 1 game):
    Row 1: Player names (columns A, B, C, ...)
    Row 2: Total scores
    Rows 3+: Per-round scores (0 = bust)
    "Winner" text in bottom cell of winning player's column

Example:
    python import_games.py games.xlsx abc-123-def-456 import_games.sql
"""

import openpyxl
import uuid
from datetime import datetime
import random
import string
import sys
import os


def generate_join_code():
    """Generate a 6-character join code (A-Z, 0-9)"""
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))


def escape_sql_string(value):
    """Escape single quotes for SQL"""
    if value is None:
        return ''
    return str(value).replace("'", "''")


def parse_game_sheet(sheet):
    """
    Parse a single Excel sheet into game data.

    Expected format:
        Row 1: Player names
        Row 2: Total scores
        Row 3+: Per-round scores
        "Winner" in bottom cell of winner's column
    """
    players = []

    # Find columns with player names (row 1)
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value and str(cell_value).strip():
            name = str(cell_value).strip()

            # Get total score from row 2
            total_cell = sheet.cell(row=2, column=col).value
            try:
                total_score = int(total_cell) if total_cell else 0
            except (ValueError, TypeError):
                total_score = 0

            players.append({
                'name': name,
                'column': col,
                'total_score': total_score,
                'turns': [],
                'is_winner': False
            })

    if not players:
        return None

    # Parse turns (row 3 onwards) and find winner marker
    for row in range(3, sheet.max_row + 2):
        for player in players:
            cell_value = sheet.cell(row=row, column=player['column']).value

            if cell_value is None or str(cell_value).strip() == '':
                continue

            cell_str = str(cell_value).strip()

            # Check for winner marker
            if cell_str.lower() == 'winner':
                player['is_winner'] = True
            else:
                # Try to parse as score
                try:
                    score = int(cell_value)
                    turn_number = len(player['turns']) + 1
                    player['turns'].append({
                        'turn_number': turn_number,
                        'score': score,
                        'is_bust': score == 0
                    })
                except (ValueError, TypeError):
                    # Skip non-numeric, non-winner cells
                    pass

    return players


def generate_sql(excel_path, creator_user_id, output_path):
    """Generate SQL INSERT statements from Excel file"""

    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    lines = []
    lines.append("-- ============================================================")
    lines.append("-- 10K Scorekeeper - Game Import SQL")
    lines.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"-- Source: {os.path.basename(excel_path)}")
    lines.append("-- ============================================================")
    lines.append("--")
    lines.append("-- Run this in Supabase SQL Editor")
    lines.append("-- All players are created as guests (is_guest=true, user_id=NULL)")
    lines.append("-- Use database/manual/claim_guest_player.sql to link accounts later")
    lines.append("--")
    lines.append("-- ============================================================")
    lines.append("")

    game_count = 0
    total_players = 0
    total_turns = 0
    player_names_seen = set()

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        players = parse_game_sheet(sheet)

        if not players:
            lines.append(f"-- Skipped sheet '{sheet_name}' (no valid data)")
            lines.append("")
            continue

        game_count += 1
        game_id = str(uuid.uuid4())
        join_code = generate_join_code()
        total_rounds = max(len(p['turns']) for p in players) if players else 0

        # Track unique player names
        for p in players:
            player_names_seen.add(p['name'])

        # Find winner
        winner = next((p for p in players if p['is_winner']), None)

        lines.append(f"-- ============================================================")
        lines.append(f"-- GAME {game_count}: {sheet_name}")
        lines.append(f"-- Players: {', '.join(p['name'] for p in players)}")
        lines.append(f"-- Rounds: {total_rounds}")
        if winner:
            lines.append(f"-- Winner: {winner['name']} ({winner['total_score']} pts)")
        else:
            lines.append(f"-- Winner: Not specified")
        lines.append(f"-- ============================================================")
        lines.append("")

        # Generate player IDs
        for i, player in enumerate(players):
            player['id'] = str(uuid.uuid4())

        winner_id = winner['id'] if winner else None
        winner_score = winner['total_score'] if winner else None

        # INSERT game as 'active' first (trigger blocks updates to 'ended' games)
        lines.append("-- Game record (insert as active, will finalize after players)")
        lines.append(f"""INSERT INTO games (
    id, created_by_user_id, join_code, status, total_rounds,
    winning_player_id, winning_score, created_at, updated_at, finished_at
) VALUES (
    '{game_id}',
    '{creator_user_id}',
    '{join_code}',
    'active',
    {total_rounds},
    NULL,
    NULL,
    NOW(), NOW(), NOW()
);""")
        lines.append("")

        # INSERT game_players
        # First player (column A) is the game creator - link to their account
        lines.append("-- Players (first player = game creator)")
        for i, player in enumerate(players):
            escaped_name = escape_sql_string(player['name'])
            is_creator = (i == 0)  # First player is the creator
            user_id = f"'{creator_user_id}'" if is_creator else "NULL"
            is_guest = "false" if is_creator else "true"

            lines.append(f"""INSERT INTO game_players (
    id, game_id, user_id, player_name, is_guest, is_on_board, total_score, display_order, created_at
) VALUES (
    '{player['id']}', '{game_id}', {user_id}, '{escaped_name}', {is_guest}, true, {player['total_score']}, {i + 1}, NOW()
);""")
        lines.append("")

        # INSERT turns
        if any(p['turns'] for p in players):
            lines.append("-- Turns")
            for player in players:
                for turn in player['turns']:
                    turn_id = str(uuid.uuid4())
                    lines.append(f"""INSERT INTO turns (
    id, game_id, player_id, turn_number, score, is_bust, is_closed, created_at
) VALUES (
    '{turn_id}', '{game_id}', '{player['id']}', {turn['turn_number']}, {turn['score']}, {str(turn['is_bust']).lower()}, true, NOW()
);""")
                    total_turns += 1
            lines.append("")

        # UPDATE game with winner and finalize status (now that players exist)
        lines.append("-- Finalize game (set winner and status)")
        if winner_id:
            lines.append(f"""UPDATE games SET
    status = 'ended',
    winning_player_id = '{winner_id}',
    winning_score = {winner_score if winner_score is not None else 'NULL'}
WHERE id = '{game_id}';""")
        else:
            lines.append(f"""UPDATE games SET status = 'ended' WHERE id = '{game_id}';""")
        lines.append("")

        total_players += len(players)

    # Summary
    lines.append("-- ============================================================")
    lines.append(f"-- IMPORT SUMMARY")
    lines.append(f"-- ============================================================")
    lines.append(f"-- Games imported: {game_count}")
    lines.append(f"-- Total player records: {total_players}")
    lines.append(f"-- Total turn records: {total_turns}")
    lines.append(f"-- Unique player names: {len(player_names_seen)}")
    if player_names_seen:
        lines.append(f"-- Players: {', '.join(sorted(player_names_seen))}")
    lines.append("-- ============================================================")
    lines.append("")
    lines.append("-- To claim guest players after they sign up, use:")
    lines.append("-- database/manual/claim_guest_player.sql")
    lines.append("")

    sql_content = '\n'.join(lines)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(sql_content)

    print(f"")
    print(f"Success! Generated: {output_path}")
    print(f"")
    print(f"  Games:   {game_count}")
    print(f"  Players: {total_players} records ({len(player_names_seen)} unique names)")
    print(f"  Turns:   {total_turns}")
    print(f"")
    print(f"Next steps:")
    print(f"  1. Open Supabase SQL Editor")
    print(f"  2. Paste contents of {output_path}")
    print(f"  3. Run the SQL")
    print(f"")

    return sql_content


def main():
    if len(sys.argv) < 3:
        print("")
        print("10K Scorekeeper - Excel to SQL Import")
        print("=" * 50)
        print("")
        print("Usage:")
        print("  python import_games.py <excel_file> <your_user_id> [output_file]")
        print("")
        print("Arguments:")
        print("  excel_file   - Path to Excel file (.xlsx) with game data")
        print("  your_user_id - Your Supabase user UUID (from profiles table)")
        print("  output_file  - Output SQL file (default: import_games.sql)")
        print("")
        print("Excel format (each sheet = 1 game):")
        print("  Row 1: Player names (column A = YOU, the game creator)")
        print("  Row 2: Total scores")
        print("  Row 3+: Per-round scores (0 = bust)")
        print("  'Winner' text in bottom cell of winner's column")
        print("")
        print("Note: Column A player is linked to your account (is_guest=false)")
        print("      Other players are guests (claimable later)")
        print("")
        print("Example:")
        print("  python import_games.py games.xlsx 123e4567-e89b-12d3-a456-426614174000")
        print("")
        sys.exit(1)

    excel_path = sys.argv[1]
    creator_id = sys.argv[2]
    output_path = sys.argv[3] if len(sys.argv) > 3 else 'import_games.sql'

    # Validate UUID format (basic check)
    try:
        uuid.UUID(creator_id)
    except ValueError:
        print(f"Error: Invalid UUID format: {creator_id}")
        print("Your user ID should look like: 123e4567-e89b-12d3-a456-426614174000")
        sys.exit(1)

    generate_sql(excel_path, creator_id, output_path)


if __name__ == '__main__':
    main()
